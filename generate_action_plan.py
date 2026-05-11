from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

wb = Workbook()

# ─── COLORS ───
NAVY = "0F172A"
BLUE = "0EA5E9"
TEAL = "0D9488"
GOLD = "F59E0B"
WHITE = "FFFFFF"
LIGHT = "F0F9FF"
LIGHT2 = "E0F2FE"
MUTED = "94A3B8"
GREEN = "22C55E"
RED = "EF4444"
PURPLE = "7C3AED"

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def cell_font(size=10, bold=False, color="1E293B"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(color):
    return PatternFill("solid", fgColor=color)

def border(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def apply_row(ws, row, values, font, fill_color=None, align=None, height=None):
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = font
        if fill_color:
            cell.fill = fill(fill_color)
        if align:
            cell.alignment = align
        cell.border = border()
    if height:
        ws.row_dimensions[row].height = height

# ══════════════════════════════════════════════════════
# SHEET 1: 90-DAY ACTION PLAN
# ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "90-Day Action Plan"
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:G1")
ws1["A1"] = "AI EXPLAINER CHANNEL — 90-Day Action Plan"
ws1["A1"].font = Font(name="Arial", size=18, bold=True, color=WHITE)
ws1["A1"].fill = fill(NAVY)
ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:G2")
ws1["A2"] = "Track your progress week by week. Mark Done when completed."
ws1["A2"].font = Font(name="Arial", size=10, italic=True, color=MUTED)
ws1["A2"].fill = fill("0C1A2E")
ws1["A2"].alignment = center()
ws1.row_dimensions[2].height = 20

# Headers
headers = ["Phase", "Week", "Task", "Category", "Priority", "Status", "Notes"]
ws1.row_dimensions[3].height = 24
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=i, value=h)
    cell.font = hdr_font(11)
    cell.fill = fill("1D4ED8")
    cell.alignment = center()
    cell.border = border()

tasks = [
    # Phase 1: Foundation (Days 1-7)
    ("Foundation\n(Days 1-7)", "Week 1", "Create YouTube channel with optimized name & description", "Setup", "HIGH", "To Do", "Channel name ideas: AI Simply, Decoded AI, Plain AI"),
    ("Foundation\n(Days 1-7)", "Week 1", "Design channel art & profile picture in Canva", "Branding", "HIGH", "To Do", "Use Canva free template — dark bg + bold channel name"),
    ("Foundation\n(Days 1-7)", "Week 1", "Create Instagram account matching YouTube brand", "Setup", "HIGH", "To Do", "Same username as YouTube if available"),
    ("Foundation\n(Days 1-7)", "Week 1", "Sign up for affiliate programs (Jasper, Copy.ai, Notion)", "Monetization", "HIGH", "To Do", "Get your unique affiliate links before first video"),
    ("Foundation\n(Days 1-7)", "Week 1", "Research top 10 AI keywords using TubeBuddy (free tier)", "SEO", "MEDIUM", "To Do", "Target keywords with high search + low competition"),
    ("Foundation\n(Days 1-7)", "Week 1", "Script first 3 videos using Claude/ChatGPT", "Content", "HIGH", "To Do", "Use prompt: 'Script 7-min explainer on [topic] for non-technical people'"),
    ("Foundation\n(Days 1-7)", "Week 1", "Record channel trailer (60-90 seconds)", "Content", "HIGH", "To Do", "Hook: 'If AI feels overwhelming, this channel is for you'"),
    ("Foundation\n(Days 1-7)", "Week 1", "Set up free ConvertKit/Beehiiv account for email list", "Setup", "MEDIUM", "To Do", "Start building email list from day 1"),
    # Phase 2: Launch (Days 8-30)
    ("Launch\n(Days 8-30)", "Week 2", "Publish Video 1: 'How ChatGPT Actually Works (Simply Explained)'", "Content", "HIGH", "To Do", "Aim for 8-10 min. Add timestamps, SEO description, affiliate links"),
    ("Launch\n(Days 8-30)", "Week 2", "Create Instagram Reel from Video 1 highlight (60-90 sec)", "Content", "HIGH", "To Do", "Hook in first 3 seconds. Add captions. Post same day as YT upload"),
    ("Launch\n(Days 8-30)", "Week 2", "Design Video 1 thumbnail in Canva", "Content", "HIGH", "To Do", "Formula: your face + surprised expression + bold 2-3 word text"),
    ("Launch\n(Days 8-30)", "Week 2", "Comment on 10 videos in your niche to build visibility", "Growth", "MEDIUM", "To Do", "Add genuine value. Never spam. Builds community visibility"),
    ("Launch\n(Days 8-30)", "Week 3", "Publish Video 2: 'I Used AI for Everything for 7 Days'", "Content", "HIGH", "To Do", "Tutorial format works very well. Include screen recordings"),
    ("Launch\n(Days 8-30)", "Week 3", "Create Instagram carousel: '5 AI Tools You Should Try'", "Content", "MEDIUM", "To Do", "Carousels get 3x more reach than static posts on IG"),
    ("Launch\n(Days 8-30)", "Week 3", "Set up Gumroad account and create first digital product", "Monetization", "MEDIUM", "To Do", "'The 50 Best ChatGPT Prompts for Beginners' — sell for $9"),
    ("Launch\n(Days 8-30)", "Week 4", "Publish Video 3: 'What is Machine Learning? (With Examples)'", "Content", "HIGH", "To Do", "Concept explainer. Use analogies. Animated visuals if possible"),
    ("Launch\n(Days 8-30)", "Week 4", "Analyze Month 1 analytics — what performed best?", "Analytics", "HIGH", "To Do", "Focus: watch time %, click-through rate, top search terms"),
    ("Launch\n(Days 8-30)", "Week 4", "Write Month 1 newsletter issue (repurpose video script)", "Content", "LOW", "To Do", "Substack or Beehiiv. Short + helpful. Link to your latest video"),
    # Phase 3: Momentum (Days 31-60)
    ("Momentum\n(Days 31-60)", "Week 5", "Publish Video 4: 'The AI Tool That Changed How I Work'", "Content", "HIGH", "To Do", "Tool deep dive. Compare 2-3 tools. Include affiliate links"),
    ("Momentum\n(Days 31-60)", "Week 5", "Set up Repurpose.io for auto-distribution", "Automation", "HIGH", "To Do", "Connect YouTube → Instagram + TikTok auto-posting"),
    ("Momentum\n(Days 31-60)", "Week 6", "Publish Video 5: 'What is Generative AI? (Easy Explanation)'", "Content", "HIGH", "To Do", "High search volume topic. Great for new subscriber traffic"),
    ("Momentum\n(Days 31-60)", "Week 6", "Reach out to 3 small AI companies about micro-sponsorships", "Monetization", "MEDIUM", "To Do", "Email template: introduce yourself, share channel stats + audience"),
    ("Momentum\n(Days 31-60)", "Week 7", "Publish Video 6: 'How AI is Changing [Industry]'", "Content", "HIGH", "To Do", "Impact story format. Human angle. Great for shares"),
    ("Momentum\n(Days 31-60)", "Week 7", "Create second digital product based on audience comments", "Monetization", "MEDIUM", "To Do", "What are people asking in your comments? Build a product for that"),
    ("Momentum\n(Days 31-60)", "Week 8", "Publish Video 7: YouTube Shorts test — repurpose best moment", "Content", "MEDIUM", "To Do", "3-5 Shorts/week can dramatically accelerate subscriber growth"),
    # Phase 4: Scale (Days 61-90)
    ("Scale\n(Days 61-90)", "Week 9-10", "Publish 2 videos per week consistently", "Content", "HIGH", "To Do", "Consistency is the #1 growth factor. Don't break the streak"),
    ("Scale\n(Days 61-90)", "Week 9-10", "Target 500 subscribers milestone", "Growth", "HIGH", "To Do", "500 subs → you can have brand deal conversations"),
    ("Scale\n(Days 61-90)", "Week 11-12", "Review analytics and double down on best content format", "Analytics", "HIGH", "To Do", "Which pillar has the highest watch time? Make more of that"),
    ("Scale\n(Days 61-90)", "Week 11-12", "Start building toward 1,000 subs + 4,000 watch hours", "Growth", "HIGH", "To Do", "YT Partner Program milestone — keep this as your 6-month goal"),
    ("Scale\n(Days 61-90)", "Week 11-12", "Refresh SEO on top 3 videos — update descriptions and tags", "SEO", "MEDIUM", "To Do", "Old videos can get new traffic with updated optimization"),
]

phase_colors = {
    "Foundation\n(Days 1-7)": "1D4ED8",
    "Launch\n(Days 8-30)": TEAL,
    "Momentum\n(Days 31-60)": PURPLE,
    "Scale\n(Days 61-90)": "DC2626",
}
priority_colors = {"HIGH": "FEE2E2", "MEDIUM": "FEF9C3", "LOW": "F0FDF4"}

for i, task in enumerate(tasks):
    row = i + 4
    row_bg = "FFFFFF" if i % 2 == 0 else "F8FAFC"
    ws1.row_dimensions[row].height = 40

    phase, week, desc, cat, priority, status, notes = task
    values = [phase, week, desc, cat, priority, status, notes]
    for j, val in enumerate(values, 1):
        cell = ws1.cell(row=row, column=j, value=val)
        cell.border = border()
        cell.alignment = left()

        if j == 1:  # Phase
            cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            cell.fill = fill(phase_colors.get(phase, NAVY))
            cell.alignment = center(wrap=True)
        elif j == 5:  # Priority
            cell.font = Font(name="Arial", size=10, bold=True, color="1E293B")
            cell.fill = fill(priority_colors.get(priority, "FFFFFF"))
            cell.alignment = center()
        elif j == 6:  # Status
            cell.font = Font(name="Arial", size=10, color="374151")
            cell.fill = fill("F3F4F6")
            cell.alignment = center()
        elif j == 3:  # Task description
            cell.font = Font(name="Arial", size=10, bold=True, color="1E293B")
            cell.fill = fill(row_bg)
        else:
            cell.font = Font(name="Arial", size=10, color="374151")
            cell.fill = fill(row_bg)

# Column widths
ws1.column_dimensions["A"].width = 16
ws1.column_dimensions["B"].width = 12
ws1.column_dimensions["C"].width = 52
ws1.column_dimensions["D"].width = 14
ws1.column_dimensions["E"].width = 11
ws1.column_dimensions["F"].width = 11
ws1.column_dimensions["G"].width = 45

ws1.freeze_panes = "A4"

# ══════════════════════════════════════════════════════
# SHEET 2: CONTENT CALENDAR
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet("Content Calendar")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:H1")
ws2["A1"] = "CONTENT CALENDAR — Video Ideas & SEO Keyword Tracker"
ws2["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws2["A1"].fill = fill("0C2D48")
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 32

ws2.merge_cells("A2:H2")
ws2["A2"] = "Use this to plan your first 20 videos. Score = estimated opportunity (1-10). Sort by Score to prioritize."
ws2["A2"].font = Font(name="Arial", size=10, italic=True, color=MUTED)
ws2["A2"].fill = fill("0F172A")
ws2["A2"].alignment = center()
ws2.row_dimensions[2].height = 18

headers2 = ["#", "Video Title Idea", "Pillar", "Target Keyword", "Est. Monthly Searches", "Competition", "Opportunity Score", "Status"]
ws2.row_dimensions[3].height = 24
for i, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = hdr_font(11)
    cell.fill = fill("0D9488")
    cell.alignment = center()
    cell.border = border()

videos = [
    (1, "How ChatGPT Actually Works (Simple Explanation)", "Concept", "how does chatgpt work", "90K+", "MEDIUM", 9, "Script Ready"),
    (2, "What is Machine Learning? (With Everyday Examples)", "Concept", "what is machine learning explained", "74K+", "MEDIUM", 9, "To Script"),
    (3, "I Used AI for Everything for 7 Days — Here's What Happened", "Tool", "using ai tools for a week", "40K+", "LOW", 9, "To Script"),
    (4, "What is a Neural Network? (Animated Explanation)", "Concept", "neural network explained simply", "60K+", "MEDIUM", 8, "To Script"),
    (5, "The Best Free AI Tools of 2026 (Ranked)", "Tool", "best free ai tools 2026", "50K+", "LOW", 9, "To Script"),
    (6, "GPT vs Gemini vs Claude — Which AI is Actually Better?", "Tool", "chatgpt vs gemini vs claude", "80K+", "HIGH", 7, "To Script"),
    (7, "How AI Actually Learns: Training Explained Simply", "Concept", "how does ai learn", "55K+", "LOW", 9, "To Script"),
    (8, "Midjourney for Complete Beginners (Make AI Images Now)", "Tool", "midjourney beginner tutorial", "45K+", "MEDIUM", 8, "To Script"),
    (9, "How AI is Changing Healthcare (What Doctors Are Saying)", "Impact", "ai in healthcare explained", "35K+", "LOW", 9, "To Script"),
    (10, "What is Generative AI? (The Simple Version)", "Concept", "generative ai explained simply", "65K+", "LOW", 9, "To Script"),
    (11, "Perplexity AI vs Google — Why I Switched", "Tool", "perplexity ai vs google", "30K+", "LOW", 9, "To Script"),
    (12, "How AI Writes Text — What's Actually Happening?", "Concept", "how does ai generate text", "40K+", "LOW", 8, "To Script"),
    (13, "The AI Tools Teachers Are Using Right Now", "Impact", "ai tools for teachers", "25K+", "LOW", 9, "To Script"),
    (14, "What is Prompt Engineering? (And Why It Matters)", "Concept", "what is prompt engineering", "55K+", "MEDIUM", 8, "To Script"),
    (15, "AI That Actually Saves Time: My Honest Workflow", "Tool", "ai productivity tools 2026", "30K+", "LOW", 8, "To Script"),
    (16, "How Stable Diffusion Works (No PhD Required)", "Concept", "how does stable diffusion work", "28K+", "LOW", 8, "To Script"),
    (17, "AI is Changing These 5 Jobs Forever", "Impact", "how ai is changing jobs", "70K+", "MEDIUM", 7, "To Script"),
    (18, "What is the Difference Between AI and ML?", "Concept", "ai vs machine learning difference", "45K+", "LOW", 9, "To Script"),
    (19, "I Built an AI Chatbot With No Coding — Here's How", "Tool", "build ai chatbot no code", "35K+", "LOW", 9, "To Script"),
    (20, "The Future of AI: What's Actually Coming Next?", "Concept", "future of ai explained", "60K+", "MEDIUM", 7, "To Script"),
]

pillar_colors = {"Concept": "EFF6FF", "Tool": "F0FDF4", "Impact": "FDF4FF"}
score_colors = {9: "DCFCE7", 8: "FEF9C3", 7: "FEE2E2"}

for i, video in enumerate(videos):
    row = i + 4
    ws2.row_dimensions[row].height = 32
    row_bg = "FFFFFF" if i % 2 == 0 else "F8FAFC"

    for j, val in enumerate(video, 1):
        cell = ws2.cell(row=row, column=j, value=val)
        cell.border = border()
        cell.alignment = left()

        if j == 1:
            cell.font = Font(name="Arial", size=11, bold=True, color=WHITE)
            cell.fill = fill(NAVY)
            cell.alignment = center()
        elif j == 2:
            cell.font = Font(name="Arial", size=10, bold=True, color="1E293B")
            cell.fill = fill(row_bg)
        elif j == 3:
            pillar = video[2]
            cell.font = Font(name="Arial", size=10, color="374151")
            cell.fill = fill(pillar_colors.get(pillar, row_bg))
            cell.alignment = center()
        elif j == 7:
            score = video[6]
            cell.font = Font(name="Arial", size=11, bold=True, color="1E293B")
            cell.fill = fill(score_colors.get(score, "FFFFFF"))
            cell.alignment = center()
        elif j == 6:
            comp = video[5]
            comp_colors = {"LOW": "DCFCE7", "MEDIUM": "FEF9C3", "HIGH": "FEE2E2"}
            cell.font = Font(name="Arial", size=10, color="374151")
            cell.fill = fill(comp_colors.get(comp, row_bg))
            cell.alignment = center()
        elif j == 8:
            cell.font = Font(name="Arial", size=10, color="374151")
            cell.fill = fill("FFF7ED" if val == "Script Ready" else row_bg)
            cell.alignment = center()
        else:
            cell.font = Font(name="Arial", size=10, color="374151")
            cell.fill = fill(row_bg)
            cell.alignment = center()

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 54
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 34
ws2.column_dimensions["E"].width = 20
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 18
ws2.column_dimensions["H"].width = 14
ws2.freeze_panes = "A4"

# ══════════════════════════════════════════════════════
# SHEET 3: MONETIZATION TRACKER
# ══════════════════════════════════════════════════════
ws3 = wb.create_sheet("Monetization Tracker")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F1")
ws3["A1"] = "MONETIZATION TRACKER — Monthly Income Overview"
ws3["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws3["A1"].fill = fill(TEAL)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 32

# Affiliate section
ws3.merge_cells("A3:F3")
ws3["A3"] = "AFFILIATE PROGRAMS"
ws3["A3"].font = Font(name="Arial", size=12, bold=True, color=WHITE)
ws3["A3"].fill = fill("0D9488")
ws3["A3"].alignment = center()
ws3.row_dimensions[3].height = 22

af_headers = ["Program", "Commission Rate", "Payout Model", "Your Link Status", "This Month Clicks", "This Month Earnings ($)"]
ws3.row_dimensions[4].height = 22
for i, h in enumerate(af_headers, 1):
    cell = ws3.cell(row=4, column=i, value=h)
    cell.font = hdr_font(10, color="FFFFFF")
    cell.fill = fill("0F766E")
    cell.alignment = center(wrap=True)
    cell.border = border()

affiliates = [
    ("Jasper AI", "30% recurring", "Monthly", "Not signed up", 0, 0),
    ("Copy.ai", "45% first year", "Monthly", "Not signed up", 0, 0),
    ("Descript", "15% recurring", "Monthly", "Not signed up", 0, 0),
    ("Notion", "$10/Pro user", "Per sale", "Not signed up", 0, 0),
    ("ConvertKit", "30% recurring", "Monthly", "Not signed up", 0, 0),
    ("Midjourney", "Varies", "Per referral", "Not signed up", 0, 0),
]

for i, af in enumerate(affiliates):
    row = i + 5
    ws3.row_dimensions[row].height = 22
    bg = "FFFFFF" if i % 2 == 0 else "F0FDFA"
    for j, val in enumerate(af, 1):
        cell = ws3.cell(row=row, column=j, value=val)
        cell.border = border()
        cell.alignment = left() if j <= 4 else center()
        cell.font = Font(name="Arial", size=10, color="1E293B")
        cell.fill = fill(bg)
        if j == 6:
            cell.number_format = '$#,##0.00'

# Total row
total_row = len(affiliates) + 5
ws3.row_dimensions[total_row].height = 24
ws3.merge_cells(f"A{total_row}:E{total_row}")
ws3[f"A{total_row}"] = "TOTAL AFFILIATE EARNINGS"
ws3[f"A{total_row}"].font = Font(name="Arial", size=11, bold=True, color=WHITE)
ws3[f"A{total_row}"].fill = fill("0D9488")
ws3[f"A{total_row}"].alignment = center()
ws3[f"A{total_row}"].border = border()
ws3[f"F{total_row}"] = f"=SUM(F5:F{total_row-1})"
ws3[f"F{total_row}"].font = Font(name="Arial", size=11, bold=True, color=WHITE)
ws3[f"F{total_row}"].fill = fill("0D9488")
ws3[f"F{total_row}"].alignment = center()
ws3[f"F{total_row}"].border = border()
ws3[f"F{total_row}"].number_format = '$#,##0.00'

# Digital Products section
prod_start = total_row + 2
ws3.merge_cells(f"A{prod_start}:F{prod_start}")
ws3[f"A{prod_start}"] = "DIGITAL PRODUCTS"
ws3[f"A{prod_start}"].font = Font(name="Arial", size=12, bold=True, color=WHITE)
ws3[f"A{prod_start}"].fill = fill(BLUE)
ws3[f"A{prod_start}"].alignment = center()
ws3.row_dimensions[prod_start].height = 22

prod_headers = ["Product Name", "Platform", "Price ($)", "Units Sold", "Revenue ($)", "Status"]
for i, h in enumerate(prod_headers, 1):
    cell = ws3.cell(row=prod_start+1, column=i, value=h)
    cell.font = hdr_font(10)
    cell.fill = fill("1D4ED8")
    cell.alignment = center()
    cell.border = border()

products = [
    ("AI Prompt Pack for Beginners (50 prompts)", "Gumroad", 9, 0, "=C{}*D{}".format(prod_start+2, prod_start+2), "Not Created"),
    ("AI Productivity Notion Template", "Gumroad", 19, 0, "=C{}*D{}".format(prod_start+3, prod_start+3), "Not Created"),
    ("AI Starter Guide PDF", "Gumroad", 7, 0, "=C{}*D{}".format(prod_start+4, prod_start+4), "Not Created"),
]

for i, prod in enumerate(products):
    row = prod_start + 2 + i
    ws3.row_dimensions[row].height = 22
    bg = "FFFFFF" if i % 2 == 0 else "EFF6FF"
    name, platform, price, units, formula, status = prod
    for j, val in enumerate([name, platform, price, units, formula, status], 1):
        cell = ws3.cell(row=row, column=j, value=val)
        cell.border = border()
        cell.alignment = left() if j == 1 else center()
        cell.font = Font(name="Arial", size=10, color="1E293B")
        cell.fill = fill(bg)
        if j in (3, 5):
            cell.number_format = '$#,##0.00'

# Income summary
sum_start = prod_start + len(products) + 3
ws3.merge_cells(f"A{sum_start}:F{sum_start}")
ws3[f"A{sum_start}"] = "MONTHLY INCOME SUMMARY"
ws3[f"A{sum_start}"].font = Font(name="Arial", size=12, bold=True, color=WHITE)
ws3[f"A{sum_start}"].fill = fill(NAVY)
ws3[f"A{sum_start}"].alignment = center()
ws3.row_dimensions[sum_start].height = 26

months = ["Month 1", "Month 2", "Month 3", "Month 4", "Month 5", "Month 6"]
sum_headers = ["Month", "Affiliate ($)", "Products ($)", "Sponsorships ($)", "YouTube Ads ($)", "Total ($)"]
for i, h in enumerate(sum_headers, 1):
    cell = ws3.cell(row=sum_start+1, column=i, value=h)
    cell.font = hdr_font(10)
    cell.fill = fill("1E3A5F")
    cell.alignment = center()
    cell.border = border()
ws3.row_dimensions[sum_start+1].height = 22

for i, month in enumerate(months):
    row = sum_start + 2 + i
    ws3.row_dimensions[row].height = 24
    bg = "FFFFFF" if i % 2 == 0 else "F8FAFC"
    ws3.cell(row=row, column=1, value=month).font = Font(name="Arial", size=10, bold=True, color="1E293B")
    ws3.cell(row=row, column=1).fill = fill(bg)
    ws3.cell(row=row, column=1).border = border()
    ws3.cell(row=row, column=1).alignment = center()
    for j in range(2, 6):
        cell = ws3.cell(row=row, column=j, value=0)
        cell.font = Font(name="Arial", size=10, color="374151")
        cell.fill = fill(bg)
        cell.border = border()
        cell.number_format = '$#,##0.00'
        cell.alignment = center()
    total_cell = ws3.cell(row=row, column=6, value=f"=SUM(B{row}:E{row})")
    total_cell.font = Font(name="Arial", size=10, bold=True, color="1E293B")
    total_cell.fill = fill("DCFCE7" if i >= 2 else bg)
    total_cell.border = border()
    total_cell.number_format = '$#,##0.00'
    total_cell.alignment = center()

ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 18
ws3.column_dimensions["F"].width = 16

# ══════════════════════════════════════════════════════
# SHEET 4: CHANNEL ANALYTICS DASHBOARD
# ══════════════════════════════════════════════════════
ws4 = wb.create_sheet("Analytics Dashboard")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:H1")
ws4["A1"] = "CHANNEL ANALYTICS DASHBOARD — Weekly Tracker"
ws4["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws4["A1"].fill = fill(PURPLE)
ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 32

ws4.merge_cells("A2:H2")
ws4["A2"] = "Fill this in every Sunday. Track the metrics that actually matter for growth."
ws4["A2"].font = Font(name="Arial", size=10, italic=True, color=MUTED)
ws4["A2"].fill = fill("1E1035")
ws4["A2"].alignment = center()
ws4.row_dimensions[2].height = 18

an_headers = ["Week", "Date", "Subscribers", "Sub Growth", "Total Views", "Watch Hours", "Top Video", "CTR %"]
ws4.row_dimensions[3].height = 24
for i, h in enumerate(an_headers, 1):
    cell = ws4.cell(row=3, column=i, value=h)
    cell.font = hdr_font(11)
    cell.fill = fill(PURPLE)
    cell.alignment = center()
    cell.border = border()

for i in range(1, 13):
    row = i + 3
    ws4.row_dimensions[row].height = 26
    bg = "FFFFFF" if i % 2 == 0 else "FAF5FF"
    ws4.cell(row=row, column=1, value=f"Week {i}").font = Font(name="Arial", size=10, bold=True, color=WHITE)
    ws4.cell(row=row, column=1).fill = fill(PURPLE)
    ws4.cell(row=row, column=1).alignment = center()
    ws4.cell(row=row, column=1).border = border()

    for j in range(2, 9):
        cell = ws4.cell(row=row, column=j)
        cell.font = Font(name="Arial", size=10, color="374151")
        cell.fill = fill(bg)
        cell.border = border()
        cell.alignment = center()
        if j == 4 and i > 1:
            cell.value = f"=C{row}-C{row-1}"
            cell.number_format = '+#,##0;-#,##0;-'
        elif j == 8:
            cell.number_format = '0.0%'

# Milestone tracker
ms_start = 17
ws4.merge_cells(f"A{ms_start}:H{ms_start}")
ws4[f"A{ms_start}"] = "KEY MILESTONES"
ws4[f"A{ms_start}"].font = Font(name="Arial", size=12, bold=True, color=WHITE)
ws4[f"A{ms_start}"].fill = fill(NAVY)
ws4[f"A{ms_start}"].alignment = center()
ws4.row_dimensions[ms_start].height = 24

milestones = [
    ("100 Subscribers", "First milestone — don't stop!", "In Progress"),
    ("500 Subscribers", "Brand deal conversations begin", "Not Started"),
    ("1,000 Subscribers", "YT Partner Program eligible", "Not Started"),
    ("4,000 Watch Hours", "YT Partner Program eligible", "Not Started"),
    ("First Affiliate Sale", "Monetization officially started", "Not Started"),
    ("First Digital Product Sale", "Passive income activated", "Not Started"),
    ("First Sponsorship", "You're a professional creator", "Not Started"),
    ("10,000 Subscribers", "Significant income unlock", "Not Started"),
]

ms_headers = ["Milestone", "Why It Matters", "Status"]
ws4.row_dimensions[ms_start+1].height = 22
for i, h in enumerate(ms_headers, 1):
    cell = ws4.cell(row=ms_start+1, column=i, value=h)
    cell.font = hdr_font(10)
    cell.fill = fill("1E3A5F")
    cell.alignment = center()
    cell.border = border()

for i, ms in enumerate(milestones):
    row = ms_start + 2 + i
    ws4.row_dimensions[row].height = 24
    bg = "FFFFFF" if i % 2 == 0 else "F8FAFC"
    status_colors = {"In Progress": "FEF9C3", "Not Started": "F3F4F6", "Done": "DCFCE7"}
    for j, val in enumerate(ms, 1):
        cell = ws4.cell(row=row, column=j, value=val)
        cell.border = border()
        cell.font = Font(name="Arial", size=10, bold=(j == 1), color="1E293B")
        cell.fill = fill(status_colors.get(ms[2], bg) if j == 3 else bg)
        cell.alignment = center() if j == 3 else left()

ws4.column_dimensions["A"].width = 10
ws4.column_dimensions["B"].width = 14
ws4.column_dimensions["C"].width = 16
ws4.column_dimensions["D"].width = 14
ws4.column_dimensions["E"].width = 14
ws4.column_dimensions["F"].width = 14
ws4.column_dimensions["G"].width = 32
ws4.column_dimensions["H"].width = 10

# ── Save ──
output = './outputs/AI_Channel_Action_Plan.xlsx'
wb.save(output)
print(f"XLSX saved: {output}")
